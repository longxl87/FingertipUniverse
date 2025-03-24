import numpy as np
import pandas as pd
from tqdm import tqdm
import lightgbm as lgb
import xgboost as xgb
from openpyxl import load_workbook
import logging

from matplotlib import pyplot as plt
from sklearn.metrics import roc_curve, auc
from sklearn.model_selection import KFold

from datetime import datetime
from pathlib import Path

from FingertipUniverse.file_utils import save_to_excel
from FingertipUniverse.feature_engine_utils import univariate,calc_psi,calc_ks,calc_auc
from FingertipUniverse.binning_utils import cut_bins

logger = logging.getLogger(__name__)


def prob2score(prob):
    """
    prob转化为概率分的工具
    :param prob:
    :return:
    """
    return round(550 - 60 / np.log(2) * np.log(prob / (1 - prob)), 0)

def feature_class(df, model_fea):
    """
    区分出数据集中数值特征和非数字特征
    :param df:pd.Dataframe
    :param model_fea:
    :return:
    """
    num_fea = []
    cate_fea = []
    for col in model_fea:
        if pd.api.types.is_numeric_dtype(df[col]):
            num_fea.append(col)
        else:
            cate_fea.append(col)
    return num_fea, cate_fea

def oppsite_features(df1, df2, model_feas, target):
    """
    计算两组数据中，反向的特征
    :param df1:
    :param df2:
    :param model_feas:
    :param target:
    :return:
    """
    oppsite_feas = []
    for fea in tqdm(model_feas):
        corr1 = df1[fea].corr(df1[target])
        corr2 = df2[fea].corr(df2[target])
        if corr1 * corr2 <= 0:
            oppsite_feas.append(fea)
    return oppsite_feas

def oppsite_feature_kfold(df, model_feas, target, n_splits=5, random_state=42):
    """
    计算该数据集中，随机分组后都不稳定的特征
    :param df:
    :param model_feas:
    :param target:
    :param n_splits:
    :param random_state:
    :return:
    """
    df = df.copy()
    df = df.reset_index(drop=True)
    assert df.index.is_unique, '输入数据的index有重复值，请优先确保数据格式的正确性'
    oppsite_feas = []
    kfold = KFold(n_splits=n_splits, shuffle=True, random_state=random_state)
    for idxs1, idxs2 in kfold.split(df):
        df1 = df.iloc[idxs1, :]
        df2 = df.iloc[idxs2, :]
        for x in tqdm(model_feas):
            corr1 = df1[x].corr(df1[target])
            corr2 = df2[x].corr(df2[target])
            if corr1 * corr2 <= 0:
                oppsite_feas.append(x)
    oppsite_feas = list(set(oppsite_feas))
    return oppsite_feas

def model_features(model, importance_type='gain', filter_zero=True, retType='df'):
    """
    分析模型特征及权重
    :param model:
    :param importance_type:
    :parm filter_zero:
    :parm retType:
    :return:
    """
    if isinstance(model,(lgb.Booster,lgb.LGBMRegressor,lgb.LGBMClassifier)): # 兼容lgb 的 原生和sklearn接口
        if isinstance(model,lgb.Booster):
            booster = model
        else: # 兼容lgb sklearn 接口
            booster = model.booster_
        fea_rs = pd.DataFrame(
            {"var": booster.feature_name(), "imps": booster.feature_importance(importance_type=importance_type)}
        ).sort_values("imps", ascending=False)
    elif isinstance(model, (xgb.sklearn.XGBClassifier, xgb.sklearn.XGBRegressor, xgb.core.Booster)): # 兼容xgb的原生和sklearn接口
        if isinstance(model, xgb.core.Booster):
            booster = model
        else: # 兼容sklearn接口
            booster = model.get_booster()
        fea_rs = pd.DataFrame(list(booster.get_score(importance_type='total_gain').items()),
                                  columns=["var", "imps"]).sort_values('imps', ascending=False)
    else:
        raise ValueError('输入的模型必须是 lgb.Booster,lgb.LGBMRegressor,lgb.LGBMClassifier,xgb.sklearn.XGBClassifier, xgb.sklearn.XGBRegressor, xgb.core.Booster ')

    if filter_zero:
        fea_rs = fea_rs[fea_rs['imps'] > 0]
    if retType == 'dict':
        return fea_rs.set_index('var')['imps'].to_dict()
    else:
        return fea_rs

def plot_roc_ks(y, x, data_desc=None):
    """
    绘制auc和ks
    :param y:
    :param x:
    :param data_desc:
    :return:
    """
    x = np.asarray(x)
    y = np.asarray(y)
    # 计算FPR, TPR 和 AUC
    fpr, tpr, _ = roc_curve(y, x)
    roc_auc = auc(fpr, tpr)
    # 按照预测概率排序
    sorted_indices = np.argsort(x)[::-1]
    sorted_y_true = y[sorted_indices]
    # 计算累计分布
    total_positives = np.sum(y == 1)
    total_negatives = np.sum(y == 0)
    cumulative_positives = np.cumsum(sorted_y_true == 1) / total_positives
    cumulative_negatives = np.cumsum(sorted_y_true == 0) / total_negatives
    # KS统计量
    ks_statistic = max(abs(cumulative_positives - cumulative_negatives))
    # 设置全局字体大小
    plt.rcParams.update({'font.size': 9})  # 将默认字体大小设置为10
    # 创建一个1行2列的子图布局
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9, 4))  # 减小图表尺寸
    # 绘制ROC曲线
    ax1.plot(fpr, tpr, color='darkorange', lw=2, label=f'ROC curve (AUC = {roc_auc:.2f})')
    ax1.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    ax1.set_xlim([0.0, 1.0])
    ax1.set_ylim([0.0, 1.05])
    ax1.set_xlabel('False Positive Rate')
    ax1.set_ylabel('True Positive Rate')
    ax1.set_title('Receiver Operating Characteristic (ROC)')
    ax1.legend(loc="lower right")
    # 绘制KS曲线
    ax2.plot(cumulative_positives, label='Cumulative Positives')
    ax2.plot(cumulative_negatives, label='Cumulative Negatives')
    ax2.axvline(np.argmax(abs(cumulative_positives - cumulative_negatives)), color='r', linestyle='--',
                label=f'Max KS: {ks_statistic:.2f}')
    ax2.set_xlabel('Sample Fraction (sorted by score)')
    ax2.set_ylabel('Cumulative Distribution Function')
    ax2.set_title('KS Statistic Curve')
    ax2.legend()
    if data_desc:  # 添加主标题
        plt.suptitle(f'Performance on {data_desc}', fontsize=14)  # 主标题字体大小也相应减小

    # 调整子图之间的间距
    plt.tight_layout(rect=[0, 0, 1, 1])  # rect参数确保标题不会被裁剪
    plt.show()

def feature_univariates(data_sets, feature_list, target, n_bins=10, method='freq',save=False):
    """
    基于数据集中的结果分析特征列表
    :param data_sets:
    :param feature_list:
    :param target:
    :param n_bins:
    :param method:
    :param save: 是否保存
    :return:
    """
    train, test, oot = None, None, None
    if len(data_sets) == 1:
        train, test, oot = data_sets[0], data_sets[0], data_sets[0]
    elif len(data_sets) == 2:
        train, test, oot = data_sets[0], data_sets[1], data_sets[1]
    elif len(data_sets) == 3:
        train, test, oot = data_sets[0], data_sets[1], data_sets[2]

    feature_summary = []
    feature_bininfos = {}

    for col in tqdm(feature_list):
        bins = cut_bins(train[col], n_bins=n_bins, method=method)
        train_fea_info = univariate(train[target], train[col], bins=bins)
        train_fea_info['feature'] = col
        test_fea_info = univariate(test[target], test[col], bins=bins)
        oot_fea_info = univariate(oot[target], oot[col], bins=bins)

        train_fea_info[['total', 'pct', 'brate', 'woe', 'iv']] = train_fea_info[
            ['total', 'pct', 'brate', 'woe', 'iv']].round(4)
        test_fea_info[['total', 'pct', 'brate', 'woe', 'iv']] = test_fea_info[
            ['total', 'pct', 'brate', 'woe', 'iv']].round(4)
        oot_fea_info[['total', 'pct', 'brate', 'woe', 'iv']] = oot_fea_info[
            ['total', 'pct', 'brate', 'woe', 'iv']].round(4)

        display_col = ['bin', 'bad', 'total', 'pct', 'brate', 'woe', 'iv']

        display_df = train_fea_info[['feature'] + display_col].merge(test_fea_info[display_col], on='bin',
                                                                     how='left').merge(oot_fea_info[display_col],
                                                                                       on='bin', how='left')
        display_df['bin'] = display_df['bin'].astype(str)
        display_df.columns = ['feature', 'bin', 'bad_train', 'total_train', 'pct_train', 'brate_train', 'woe_train',
                              'iv_train', 'bad_test',
                              'total_test', 'pct_test', 'brate_test', 'woe_test', 'iv_test', 'bad_oot', 'total_oot',
                              'pct_oot',
                              'brate_oot', 'woe_oot', 'iv_oot']

        iv_train = train_fea_info['iv'].max()
        iv_test = test_fea_info['iv'].max()
        iv_oot = oot_fea_info['iv'].max()
        csi1 = calc_psi(train[col], test[col])
        csi2 = calc_psi(train[col], oot[col])
        feature_info = {'feature': col,
                        'iv_train': round(iv_train, 4), 'iv_test': round(iv_test, 4), 'iv_oot': round(iv_oot, 4),
                        'csi(train-test)': round(csi1, 4), 'csi(train-oot)': round(csi2, 4)}

        feature_summary.append(feature_info)
        feature_bininfos[col] = display_df
    feature_summary_df = pd.DataFrame(feature_summary)
    feature_summary_df['ratio'] = feature_summary_df['iv_oot'] / feature_summary_df['iv_train']
    feature_summary_df['ratio'] = feature_summary_df['ratio'].round(4)
    if save:
        wb = load_workbook(Path(__file__).parent / 'model_report_template_v2.xlsx')
        feature_summary_sheet = wb["feature_summary"]
        feature_bininfo_sheet = wb["feauture_bininfo"]

        # 遍历并删除指定 Sheet
        for sheet_name in ['model_sumary','model_desc','score_bininfo']:
            if sheet_name in wb.sheetnames:  # 确保 Sheet 存在
                sheet = wb[sheet_name]
                wb.remove(sheet)

        # 写 feature_summary
        save_to_excel(feature_summary_df, feature_summary_sheet, 2, 1)

        feature_bininfos_list = [feature_bininfos[x] for x in feature_summary_df['feature'].to_list()]

        # 写 feauture_bininfo
        start_index = 3
        for feature_info in feature_bininfos_list:
            feature_info['bin'] = feature_info['bin'].astype(str)
            save_to_excel(feature_info, feature_bininfo_sheet, start_index, 1)
            start_index = start_index + len(feature_info) + 1

        time_str = datetime.now().strftime("%Y%m%d%H%M%S")
        file_name = f'feature_report{time_str}.xlsx'
        wb.save(file_name)
        logger.info(f"报告保存文件:{file_name}")
    return feature_summary_df, feature_bininfos


def model_report(data_sets, target,model_obj, time_col='loan_time',score_name='score_v0',p2score_fun=prob2score):
    """
    prob转化为概率分的工具
    :param data_sets :list[pd.Dataframe] 数据集
    :param target:str 好坏标签
    :param time_col:str 时间字段，
    :param p2score_fun:  概率转分的计算公式
    :param score_name:  模型名称
    :param model_obj :训练好的模型对象
    :return:
    """
    feature_importance_dict = model_features(model_obj, importance_type='gain', retType='dict',filter_zero=False)
    feature_list = feature_importance_dict['var'].to_list()

    assert len(data_sets) > 0 and len(data_sets) < 4, f"datas只能是1-3个数据集"

    wb = load_workbook(Path(__file__).parent / 'model_report_template_v2.xlsx')
    model_sumary_sheet = wb['model_sumary']
    model_desc_sheet = wb['model_desc']
    score_bininfo_sheet = wb['score_bininfo']
    feature_summary_sheet = wb["feature_summary"]
    feature_bininfo_sheet = wb["feauture_bininfo"]
    n_bins = 10
    method = 'freq'
    train, test, oot = None,None,None
    if len(data_sets) == 1:
        train, test, oot = data_sets[0], data_sets[0], data_sets[0]
    elif len(data_sets) == 2:
        train, test, oot = data_sets[0], data_sets[1], data_sets[1]
    elif len(data_sets) == 3:
        train, test, oot = data_sets[0], data_sets[1], data_sets[2]

    train['prob'] = model_obj.predict(train[feature_list])
    test['prob'] = model_obj.predict(test[feature_list])
    oot['prob'] = model_obj.predict(oot[feature_list])
    train['score'] = train['prob'].map(p2score_fun)
    test['score'] = test['prob'].map(p2score_fun)
    oot['score'] = oot['prob'].map(p2score_fun)

    ## 计算特征报告格式
    feature_summary_df, feature_bininfos = feature_univariates(data_sets,feature_list, target, n_bins,
                                                               method)

    if feature_importance_dict is not None:
        feature_summary_df['imps'] = feature_summary_df['feature'].map(feature_importance_dict)
    else:
        feature_summary_df['imps'] = feature_summary_df['iv_train'] + feature_summary_df['iv_test'] + \
                                     feature_summary_df['iv_oot']
    feature_summary_df = feature_summary_df[['imps', 'feature', 'iv_train', 'iv_test', 'iv_oot', 'csi(train-test)',
                                             'csi(train-oot)', 'ratio']].sort_values('imps', ascending=False)

    # 写 feature_summary
    save_to_excel(feature_summary_df, feature_summary_sheet, 2, 1)

    feature_bininfos_list = [feature_bininfos[x] for x in feature_summary_df['feature'].to_list()]

    # 写 feauture_bininfo
    start_index = 3
    for feature_info in feature_bininfos_list:
        feature_info['bin'] = feature_info['bin'].astype(str)
        save_to_excel(feature_info, feature_bininfo_sheet, start_index, 1)
        start_index = start_index + len(feature_info) + 1



    # 生成 score_bininfo 的内容
    train_df = pd.concat([ train[['score', target]] ,
                                test[['score', target]] ], axis=0)
    score_bin1 = cut_bins(train_df['score'], n_bins=10, method='freq')
    left_tbl = univariate(train_df[target],train_df['score'],  bins=score_bin1)
    right_tbl = univariate(oot[target],oot['score'],  bins=score_bin1)
    tbl1 = left_tbl.merge(right_tbl, how='left', on='bin')
    tbl1['bin'] = tbl1['bin'].astype(str)
    save_to_excel(tbl1, score_bininfo_sheet, 3, 1)

    score_bin2 = cut_bins(train_df['score'], n_bins=10, method='dist')
    left_tbl = univariate(train_df[target],train_df['score'],  bins=score_bin2)
    right_tbl = univariate(oot[target],oot['score'],  bins=score_bin2)
    tbl2 = left_tbl.merge(right_tbl, how='left', on='bin')
    tbl2['bin'] = tbl2['bin'].astype(str)
    save_to_excel(tbl2, score_bininfo_sheet, 16, 1)

    score_bin3 = cut_bins(oot['score'], n_bins=10, method='freq')
    left_tbl = univariate(oot[target],oot['score'],  bins=score_bin3)
    right_tbl = univariate(train_df[target],train_df['score'],  bins=score_bin3)
    tbl3 = left_tbl.merge(right_tbl, how='left', on='bin')
    tbl3['bin'] = tbl3['bin'].astype(str)
    save_to_excel(tbl3, score_bininfo_sheet, 29, 1)

    #  model_desc 页面
    model_desc_df = pd.DataFrame([{
        'score_name':score_name,
        'train_ks':calc_ks(train[target], train['prob']),
        'test_ks':calc_ks(test[target], test['prob']),
        'oot_ks':calc_ks(oot[target], oot['prob']),
        'auc_train': calc_auc(train[target], train['prob']),
        'auc_test': calc_auc(test[target], test['prob']),
        'auc_oot':calc_auc(oot[target], oot['prob']),
        'psi': calc_psi(train['score'], train['score']),
    }])
    save_to_excel(model_desc_df,model_desc_sheet,4,2)

    data_statis = []
    def _data_base_info(df,name):
        rs = { 'data':name ,
                 'total': df[target].count(),
                 'good': df[target].count() - df[target].sum() ,
                 'bad' : df[target].sum(),
                 'brate': df[target].mean(),
                 }
        if time_col in df.columns:
            rs['start'] = df[time_col].min()
            rs['end'] = df[time_col].max()
        data_statis.append(rs)

    _data_base_info(train,'train')
    _data_base_info(test, 'test')
    _data_base_info(oot, 'oot')

    if time_col in train.columns:
        total_df = pd.concat([train[[target,time_col]],test[[target,time_col]],oot[[target,time_col]]] ,axis=0)
    else:
        total_df = pd.concat([train[[target]], test[[target]], oot[[target]]], axis=0)
    _data_base_info(total_df, 'total')
    data_info_df = pd.DataFrame(data_statis)
    save_to_excel(data_info_df, model_sumary_sheet, 10, 2)

    score_name_df = pd.DataFrame({'model_name':[score_name]})
    save_to_excel(score_name_df, model_sumary_sheet, 5, 2)

    time_str = datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f'model_report{time_str}.xlsx'
    wb.save(file_name)
    logger.info(f"报告保存文件:{file_name}")
